"""
Emission Factor Normalization Script
=====================================
Reads raw Excel files from EPA, DEFRA, and GHG Protocol and outputs
a single normalized JSON dataset with all factors in kg CO2e per
common unit denominators.

Unit conversion constants and methodology choices are documented inline.
"""

import json
import openpyxl
from pathlib import Path
from datetime import date

RAW_DIR = Path(__file__).parent.parent / "data" / "raw"
OUT_DIR = Path(__file__).parent.parent / "data" / "normalized"

# ============================================================
# UNIT CONVERSION CONSTANTS
# ============================================================
# All conversions are sourced and documented for traceability.

LB_TO_KG = 0.45359237          # exact, by definition
MILE_TO_KM = 1.609344          # exact, by definition
MWH_TO_KWH = 1000.0
SHORT_TON_TO_KG = 907.18474   # 1 short ton = 2000 lbs
MMBTU_TO_KWH = 293.07107      # 1 mmBtu = 293.07107 kWh (thermochemical)
TJ_TO_KWH = 277777.778        # 1 TJ = 277,777.778 kWh
SCF_TO_M3 = 0.0283168         # 1 standard cubic foot = 0.0283168 m³

# GWP values (IPCC AR5, used by EPA 2025)
GWP_CH4 = 28
GWP_N2O = 265


def make_source(name, edition, publisher, url, last_updated, underlying=None):
    """Create a source metadata dict."""
    s = {
        "name": name,
        "edition": edition,
        "publisher": publisher,
        "url": url,
        "last_updated": last_updated,
    }
    if underlying:
        s["underlying_data_source"] = underlying
    return s


EPA_SOURCE = make_source(
    name="EPA GHG Emission Factors Hub",
    edition="2025",
    publisher="US Environmental Protection Agency",
    url="https://www.epa.gov/climateleadership/ghg-emission-factors-hub",
    last_updated="2025-01-15",
)

DEFRA_SOURCE = make_source(
    name="UK Government GHG Conversion Factors",
    edition="2025",
    publisher="Department for Energy Security and Net Zero (DESNZ)",
    url="https://www.gov.uk/government/publications/greenhouse-gas-reporting-conversion-factors-2025",
    last_updated="2025-06-10",
)

GHG_PROTOCOL_SOURCE = make_source(
    name="GHG Protocol Cross-Sector Emission Factors",
    edition="V2.0",
    publisher="Greenhouse Gas Protocol (WRI/WBCSD)",
    url="https://ghgprotocol.org/calculation-tools-and-guidance",
    last_updated="2024-03-13",
)


# ============================================================
# EPA EXTRACTION
# ============================================================

def extract_epa(wb):
    """Extract and normalize emission factors from EPA Excel file."""
    ws = wb["Emission Factors Hub"]
    factors = []

    # --- Natural Gas (Table 1, row 38) ---
    # EPA: CO2=53.06 kg CO2/mmBtu, CH4=1 g CH4/mmBtu, N2O=0.1 g N2O/mmBtu
    co2_per_mmbtu = _cell(ws, 38, 5)  # kg CO2 / mmBtu
    ch4_per_mmbtu = _cell(ws, 38, 6)  # g CH4 / mmBtu
    n2o_per_mmbtu = _cell(ws, 38, 7)  # g N2O / mmBtu

    if co2_per_mmbtu:
        # Convert to kg CO2e per kWh
        # 1 mmBtu = 293.07107 kWh
        ch4_co2e = (ch4_per_mmbtu / 1000) * GWP_CH4 if ch4_per_mmbtu else 0
        n2o_co2e = (n2o_per_mmbtu / 1000) * GWP_N2O if n2o_per_mmbtu else 0
        total_co2e_per_mmbtu = co2_per_mmbtu + ch4_co2e + n2o_co2e
        total_co2e_per_kwh = total_co2e_per_mmbtu / MMBTU_TO_KWH

        factors.append({
            "activity_id": "natural_gas_stationary",
            "activity_name": "Natural Gas (Stationary Combustion)",
            "category": "stationary_combustion",
            "scope": "scope_1",
            "source": EPA_SOURCE,
            "factor": {
                "value_kg_co2e": round(total_co2e_per_kwh, 6),
                "unit_denominator": "kWh",
                "co2_component": round(co2_per_mmbtu / MMBTU_TO_KWH, 6),
                "ch4_component": round(ch4_co2e / MMBTU_TO_KWH, 6),
                "n2o_component": round(n2o_co2e / MMBTU_TO_KWH, 6),
            },
            "original_value": {
                "value": co2_per_mmbtu,
                "unit": "kg CO2 per mmBtu",
                "notes": "CO2 only; CH4 and N2O added during normalization using AR5 GWPs"
            },
            "geographic_scope": "US",
            "methodology_notes": "Higher Heating Value (HHV) basis. IPCC AR5 GWPs (CH4=28, N2O=265). CH4 and N2O converted to CO2e and added.",
            "staleness_flag": False,
        })

    # --- Electricity US Average (Table 6, row 366) ---
    # Columns verified: col 5=CO2, col 6=CH4, col 7=N2O (total output)
    co2_lb_mwh = _cell(ws, 366, 5)  # lb CO2 / MWh
    ch4_lb_mwh = _cell(ws, 366, 6)  # lb CH4 / MWh
    n2o_lb_mwh = _cell(ws, 366, 7)  # lb N2O / MWh

    if co2_lb_mwh:
        # Convert lb/MWh to kg CO2e/kWh
        co2_kg_kwh = (co2_lb_mwh * LB_TO_KG) / MWH_TO_KWH
        ch4_co2e_kg_kwh = (ch4_lb_mwh * LB_TO_KG * GWP_CH4) / MWH_TO_KWH if ch4_lb_mwh else 0
        n2o_co2e_kg_kwh = (n2o_lb_mwh * LB_TO_KG * GWP_N2O) / MWH_TO_KWH if n2o_lb_mwh else 0
        total = co2_kg_kwh + ch4_co2e_kg_kwh + n2o_co2e_kg_kwh

        factors.append({
            "activity_id": "electricity_grid",
            "activity_name": "Grid Electricity",
            "category": "electricity",
            "scope": "scope_2",
            "source": EPA_SOURCE,
            "factor": {
                "value_kg_co2e": round(total, 6),
                "unit_denominator": "kWh",
                "co2_component": round(co2_kg_kwh, 6),
                "ch4_component": round(ch4_co2e_kg_kwh, 6),
                "n2o_component": round(n2o_co2e_kg_kwh, 6),
            },
            "original_value": {
                "value": co2_lb_mwh,
                "unit": "lb CO2 per MWh (total output, US average)",
                "notes": "Source: eGRID2023. Converted from lb/MWh to kg CO2e/kWh."
            },
            "geographic_scope": "US (national average)",
            "methodology_notes": "eGRID2023 total output emission factor. US national average across all subregions. CH4 and N2O converted to CO2e using AR5 GWPs.",
            "staleness_flag": False,
        })

    # --- Air Travel (Table 10, rows 513-515) ---
    air_rows = [
        (513, "air_travel_short_haul", "Air Travel - Short Haul", "< 300 miles / ~483 km"),
        (514, "air_travel_medium_haul", "Air Travel - Medium Haul", "300-2300 miles / ~483-3701 km"),
        (515, "air_travel_long_haul", "Air Travel - Long Haul", ">= 2300 miles / ~3701+ km"),
    ]
    for row, aid, aname, dist_note in air_rows:
        co2 = _cell(ws, row, 4)   # kg CO2 / passenger-mile
        ch4 = _cell(ws, row, 5)   # g CH4 / passenger-mile
        n2o = _cell(ws, row, 6)   # g N2O / passenger-mile

        if co2:
            # Convert per passenger-mile to per passenger-km
            ch4_co2e = (ch4 / 1000) * GWP_CH4 if ch4 else 0
            n2o_co2e = (n2o / 1000) * GWP_N2O if n2o else 0
            total_per_mile = co2 + ch4_co2e + n2o_co2e
            total_per_km = total_per_mile / MILE_TO_KM

            factors.append({
                "activity_id": aid,
                "activity_name": aname,
                "category": "air_travel",
                "scope": "scope_3",
                "source": EPA_SOURCE,
                "factor": {
                    "value_kg_co2e": round(total_per_km, 6),
                    "unit_denominator": "passenger-km",
                    "co2_component": round(co2 / MILE_TO_KM, 6),
                    "ch4_component": round(ch4_co2e / MILE_TO_KM, 6),
                    "n2o_component": round(n2o_co2e / MILE_TO_KM, 6),
                },
                "original_value": {
                    "value": co2,
                    "unit": "kg CO2 per passenger-mile",
                    "notes": f"Distance category: {dist_note}. CO2 only in original; CH4/N2O added during normalization."
                },
                "geographic_scope": "US",
                "methodology_notes": "Does NOT include radiative forcing. EPA sources these from 'high altitude calculations' per the footnote. Converted from passenger-miles to passenger-km.",
                "staleness_flag": False,
            })

    # --- Passenger Vehicles (Table 10, rows 504-505) ---
    vehicle_rows = [
        (504, "passenger_car", "Passenger Car (Gasoline)", "Passenger Car"),
        (505, "light_duty_truck", "Light-Duty Truck", "Light-Duty Truck"),
    ]
    for row, aid, aname, vtype in vehicle_rows:
        co2 = _cell(ws, row, 4)   # kg CO2 / vehicle-mile
        ch4 = _cell(ws, row, 5)   # g CH4 / vehicle-mile
        n2o = _cell(ws, row, 6)   # g N2O / vehicle-mile

        if co2:
            ch4_co2e = (ch4 / 1000) * GWP_CH4 if ch4 else 0
            n2o_co2e = (n2o / 1000) * GWP_N2O if n2o else 0
            total_per_mile = co2 + ch4_co2e + n2o_co2e
            total_per_km = total_per_mile / MILE_TO_KM

            factors.append({
                "activity_id": aid,
                "activity_name": aname,
                "category": "ground_transport",
                "scope": "scope_1",
                "source": EPA_SOURCE,
                "factor": {
                    "value_kg_co2e": round(total_per_km, 6),
                    "unit_denominator": "vehicle-km",
                    "co2_component": round(co2 / MILE_TO_KM, 6),
                    "ch4_component": round(ch4_co2e / MILE_TO_KM, 6),
                    "n2o_component": round(n2o_co2e / MILE_TO_KM, 6),
                },
                "original_value": {
                    "value": co2,
                    "unit": "kg CO2 per vehicle-mile",
                    "notes": f"Vehicle type: {vtype}. Converted from vehicle-miles to vehicle-km."
                },
                "geographic_scope": "US",
                "methodology_notes": "Mixed fleet average. CO2 only in original; CH4/N2O added using AR5 GWPs.",
                "staleness_flag": False,
            })

    # --- Freight - Medium/Heavy Duty Truck (Table 8, row 422) ---
    co2 = _cell(ws, 422, 4)
    ch4 = _cell(ws, 422, 5)
    n2o = _cell(ws, 422, 6)
    if co2:
        ch4_co2e = (ch4 / 1000) * GWP_CH4 if ch4 else 0
        n2o_co2e = (n2o / 1000) * GWP_N2O if n2o else 0
        total_per_mile = co2 + ch4_co2e + n2o_co2e
        total_per_km = total_per_mile / MILE_TO_KM

        factors.append({
            "activity_id": "freight_road_truck",
            "activity_name": "Freight - Road (Heavy Truck)",
            "category": "freight",
            "scope": "scope_3",
            "source": EPA_SOURCE,
            "factor": {
                "value_kg_co2e": round(total_per_km, 6),
                "unit_denominator": "vehicle-km",
                "co2_component": round(co2 / MILE_TO_KM, 6),
                "ch4_component": round(ch4_co2e / MILE_TO_KM, 6),
                "n2o_component": round(n2o_co2e / MILE_TO_KM, 6),
            },
            "original_value": {
                "value": co2,
                "unit": "kg CO2 per vehicle-mile",
                "notes": "Medium- and Heavy-Duty Truck. Converted from vehicle-miles to vehicle-km."
            },
            "geographic_scope": "US",
            "methodology_notes": "Vehicle-mile basis (not ton-mile). For weight-distance, see row 425.",
            "staleness_flag": False,
        })

    # --- Rail (Table 10, row 509) ---
    co2 = _cell(ws, 509, 4)
    ch4 = _cell(ws, 509, 5)
    n2o = _cell(ws, 509, 6)
    if co2:
        ch4_co2e = (ch4 / 1000) * GWP_CH4 if ch4 else 0
        n2o_co2e = (n2o / 1000) * GWP_N2O if n2o else 0
        total_per_mile = co2 + ch4_co2e + n2o_co2e
        total_per_km = total_per_mile / MILE_TO_KM

        factors.append({
            "activity_id": "rail_intercity",
            "activity_name": "Intercity Rail (National Average)",
            "category": "ground_transport",
            "scope": "scope_3",
            "source": EPA_SOURCE,
            "factor": {
                "value_kg_co2e": round(total_per_km, 6),
                "unit_denominator": "passenger-km",
                "co2_component": round(co2 / MILE_TO_KM, 6),
                "ch4_component": round(ch4_co2e / MILE_TO_KM, 6),
                "n2o_component": round(n2o_co2e / MILE_TO_KM, 6),
            },
            "original_value": {
                "value": co2,
                "unit": "kg CO2 per passenger-mile",
                "notes": "Intercity Rail - National Average"
            },
            "geographic_scope": "US",
            "staleness_flag": False,
        })

    # --- Bus (Table 10, row 512) ---
    co2 = _cell(ws, 512, 4)
    ch4 = _cell(ws, 512, 5)
    n2o = _cell(ws, 512, 6)
    if co2:
        ch4_co2e = (ch4 / 1000) * GWP_CH4 if ch4 else 0
        n2o_co2e = (n2o / 1000) * GWP_N2O if n2o else 0
        total_per_mile = co2 + ch4_co2e + n2o_co2e
        total_per_km = total_per_mile / MILE_TO_KM

        factors.append({
            "activity_id": "bus",
            "activity_name": "Bus",
            "category": "ground_transport",
            "scope": "scope_3",
            "source": EPA_SOURCE,
            "factor": {
                "value_kg_co2e": round(total_per_km, 6),
                "unit_denominator": "passenger-km",
            },
            "original_value": {
                "value": co2,
                "unit": "kg CO2 per passenger-mile",
            },
            "geographic_scope": "US",
            "staleness_flag": False,
        })

    return factors


# ============================================================
# DEFRA EXTRACTION
# ============================================================

def extract_defra(wb):
    """Extract and normalize emission factors from DEFRA flat file."""
    ws = wb["Factors by Category"]
    factors = []

    # Build a lookup: scan all rows, index by (scope, l1, l2, l3, ghg_unit)
    rows_by_key = {}
    for row_num in range(7, ws.max_row + 1):
        scope = _cell(ws, row_num, 2) or ""
        l1 = _cell(ws, row_num, 3) or ""
        l2 = _cell(ws, row_num, 4) or ""
        l3 = _cell(ws, row_num, 5) or ""
        uom = _cell(ws, row_num, 8) or ""
        ghg = _cell(ws, row_num, 9) or ""
        val = _cell(ws, row_num, 10)
        key = (str(scope).strip(), str(l1).strip(), str(l2).strip(), str(l3).strip())
        if key not in rows_by_key:
            rows_by_key[key] = {}
        rows_by_key[key][str(ghg).strip()] = {"value": val, "uom": str(uom).strip()}

    def get_factor(scope, l1, l2, l3):
        """Get the kg CO2e total factor for a given DEFRA path."""
        key = (scope, l1, l2, l3)
        if key in rows_by_key and "kg CO2e" in rows_by_key[key]:
            entry = rows_by_key[key]["kg CO2e"]
            return entry["value"], entry["uom"]
        return None, None

    def get_components(scope, l1, l2, l3):
        """Get CO2, CH4, N2O components if available."""
        key = (scope, l1, l2, l3)
        components = {}
        if key in rows_by_key:
            for ghg_key, entry in rows_by_key[key].items():
                if "CO2 per unit" in ghg_key:
                    components["co2"] = entry["value"]
                elif "CH4 per unit" in ghg_key:
                    components["ch4"] = entry["value"]
                elif "N2O per unit" in ghg_key:
                    components["n2o"] = entry["value"]
        return components

    # --- Natural Gas ---
    # Scan for the kWh (Net CV) variant specifically
    kwh_val = None
    for row_num in range(7, ws.max_row + 1):
        scope = str(_cell(ws, row_num, 2) or "").strip()
        l1 = str(_cell(ws, row_num, 3) or "").strip()
        l2 = str(_cell(ws, row_num, 4) or "").strip()
        l3 = str(_cell(ws, row_num, 5) or "").strip()
        ghg = str(_cell(ws, row_num, 9) or "").strip()
        r_uom = str(_cell(ws, row_num, 8) or "").strip()
        r_val = _cell(ws, row_num, 10)
        if (scope == "Scope 1" and l1 == "Fuels" and l2 == "Gaseous fuels"
            and l3 == "Natural gas" and ghg == "kg CO2e"
            and "kwh" in r_uom.lower() and isinstance(r_val, (int, float))):
            kwh_val = r_val
            break

    if True:  # Keep indentation level
        if kwh_val:
            factors.append({
                "activity_id": "natural_gas_stationary",
                "activity_name": "Natural Gas (Stationary Combustion)",
                "category": "stationary_combustion",
                "scope": "scope_1",
                "source": DEFRA_SOURCE,
                "factor": {
                    "value_kg_co2e": round(kwh_val, 6),
                    "unit_denominator": "kWh",
                },
                "original_value": {
                    "value": kwh_val,
                    "unit": "kg CO2e per kWh (Net CV)",
                    "notes": "Net Calorific Value (LHV) basis. Already in kg CO2e."
                },
                "geographic_scope": "UK",
                "methodology_notes": "Net Calorific Value (Lower Heating Value) basis. Note: EPA uses Gross/HHV which gives ~10% higher energy content per unit fuel.",
                "staleness_flag": False,
            })

    # --- Electricity ---
    val, uom = get_factor("Scope 2", "UK electricity", "Electricity generated", "Electricity: UK")
    if val:
        factors.append({
            "activity_id": "electricity_grid",
            "activity_name": "Grid Electricity",
            "category": "electricity",
            "scope": "scope_2",
            "source": DEFRA_SOURCE,
            "factor": {
                "value_kg_co2e": round(val, 6),
                "unit_denominator": "kWh",
            },
            "original_value": {
                "value": val,
                "unit": "kg CO2e per kWh",
                "notes": "UK national grid average, generation only (excludes T&D losses)"
            },
            "geographic_scope": "UK",
            "methodology_notes": "UK national grid average. T&D losses are reported separately under Scope 3. This is generation-only.",
            "staleness_flag": False,
        })

    # --- Air Travel ---
    # DEFRA has duplicate kg CO2e rows per flight: WITH and WITHOUT radiative forcing.
    # The first occurrence is WITH RF, second WITHOUT. We also need Level 4 = "Average passenger".
    # Scan manually to get the right values.
    # NOTE: DEFRA "Short-haul" (<3700km) covers a MUCH wider range than EPA "Short Haul" (<300mi/483km).
    # We give DEFRA short-haul its own activity_id to avoid misleading comparisons.
    # DEFRA long-haul (>3700km) aligns with EPA long-haul (>=2300mi/~3701km) — these are comparable.
    air_mappings = [
        ("Short-haul, to/from UK", "air_travel_short_haul_defra", "Air Travel - Short Haul (DEFRA: <3700km)", "Up to 3700 km. WARNING: Not comparable to EPA short-haul (<300 miles/483km). Covers EPA short + medium combined."),
        ("Long-haul, to/from UK", "air_travel_long_haul", "Air Travel - Long Haul", "Over 3700 km — comparable to EPA >=2300 miles"),
        ("Domestic, to/from UK", "air_travel_domestic", "Air Travel - Domestic", "Domestic flights within UK"),
        ("International, to/from non-UK", "air_travel_international", "Air Travel - International", "International, non-UK origin/destination"),
    ]
    for l3_target, aid, aname, note in air_mappings:
        # Find the FIRST kg CO2e row for this flight type with "Average passenger" in Level 4
        # The first is WITH radiative forcing
        for row_num in range(7, ws.max_row + 1):
            scope = str(_cell(ws, row_num, 2) or "").strip()
            l1 = str(_cell(ws, row_num, 3) or "").strip()
            l2 = str(_cell(ws, row_num, 4) or "").strip()
            l3 = str(_cell(ws, row_num, 5) or "").strip()
            l4 = str(_cell(ws, row_num, 6) or "").strip()
            ghg = str(_cell(ws, row_num, 9) or "").strip()
            val = _cell(ws, row_num, 10)

            if (scope == "Scope 3" and l1 == "Business travel- air" and l2 == "Flights"
                and l3 == l3_target and "average passenger" in l4.lower()
                and ghg == "kg CO2e" and val is not None and isinstance(val, (int, float))):
                factors.append({
                    "activity_id": aid,
                    "activity_name": aname,
                    "category": "air_travel",
                    "scope": "scope_3",
                    "source": DEFRA_SOURCE,
                    "factor": {
                        "value_kg_co2e": round(val, 6),
                        "unit_denominator": "passenger-km",
                    },
                    "original_value": {
                        "value": val,
                        "unit": "kg CO2e per passenger-km",
                        "notes": note,
                    },
                    "geographic_scope": "UK",
                    "methodology_notes": "Includes radiative forcing (RF) multiplier. 'Average passenger' class. DEFRA recommends including RF for company reporting. The WITHOUT-RF value is approximately 60% of this number.",
                    "staleness_flag": False,
                })
                break  # Take only the first match (WITH RF)

    # --- Passenger Vehicles ---
    car_sizes = [
        ("Average car", "passenger_car", "Passenger Car (Average)"),
        ("Small car", "passenger_car_small", "Passenger Car (Small)"),
        ("Medium car", "passenger_car_medium", "Passenger Car (Medium)"),
        ("Large car", "passenger_car_large", "Passenger Car (Large)"),
    ]
    for l3_name, aid, aname in car_sizes:
        val, uom = get_factor("Scope 1", "Passenger vehicles", "Cars (by size)", l3_name)
        if val:
            factors.append({
                "activity_id": aid,
                "activity_name": aname,
                "category": "ground_transport",
                "scope": "scope_1",
                "source": DEFRA_SOURCE,
                "factor": {
                    "value_kg_co2e": round(val, 6),
                    "unit_denominator": "vehicle-km",
                },
                "original_value": {
                    "value": val,
                    "unit": "kg CO2e per km",
                },
                "geographic_scope": "UK",
                "methodology_notes": "UK fleet average by size class. Includes petrol, diesel, hybrid, and unknown fuel types.",
                "staleness_flag": False,
            })

    # --- Hotel Stays ---
    hotel_countries = [
        ("UK", "hotel_stay_uk", "Hotel Stay (UK)"),
        ("United States", "hotel_stay_us", "Hotel Stay (US)"),
    ]
    for country, aid, aname in hotel_countries:
        val, uom = get_factor("Scope 3", "Hotel stay", "Hotel stay", country)
        if val:
            factors.append({
                "activity_id": aid,
                "activity_name": aname,
                "category": "accommodation",
                "scope": "scope_3",
                "source": DEFRA_SOURCE,
                "factor": {
                    "value_kg_co2e": round(val, 6),
                    "unit_denominator": "room-night",
                },
                "original_value": {
                    "value": val,
                    "unit": "kg CO2e per room per night",
                },
                "geographic_scope": country,
                "staleness_flag": False,
            })

    # --- Steel and Aluminum ---
    # DEFRA has multiple kg CO2e rows per material differentiated by "Column Text" (col 7):
    #   "Primary material production" = total embodied carbon (USE THIS)
    #   "Closed-loop source" = net of recycled content credit
    # We must match on Column Text to get the right value.
    mat_mappings = [
        ("Metal: steel cans", "steel", "Steel (Cans)", "Steel cans — primary material production. Not raw steel."),
        ("Metal: aluminium cans and foil (excl. forming)", "aluminum", "Aluminum (Cans/Foil)", "Aluminium cans and foil, excluding forming. Primary material production."),
    ]
    for l3_name, aid, aname, note in mat_mappings:
        for row_num in range(7, ws.max_row + 1):
            scope = str(_cell(ws, row_num, 2) or "").strip()
            l1 = str(_cell(ws, row_num, 3) or "").strip()
            l2 = str(_cell(ws, row_num, 4) or "").strip()
            l3 = str(_cell(ws, row_num, 5) or "").strip()
            col_text = str(_cell(ws, row_num, 7) or "").strip()
            ghg = str(_cell(ws, row_num, 9) or "").strip()
            val = _cell(ws, row_num, 10)
            if (scope == "Scope 3" and l1 == "Material use" and l2 == "Metal"
                and l3 == l3_name and ghg == "kg CO2e"
                and "primary material production" in col_text.lower()
                and val is not None and isinstance(val, (int, float))):
                factors.append({
                    "activity_id": aid,
                    "activity_name": aname,
                    "category": "materials",
                    "scope": "scope_3",
                    "source": DEFRA_SOURCE,
                    "factor": {
                        "value_kg_co2e": round(val, 6),
                        "unit_denominator": "tonne",
                    },
                    "original_value": {
                        "value": val,
                        "unit": "kg CO2e per tonne",
                        "notes": note,
                    },
                    "geographic_scope": "UK",
                    "methodology_notes": "Embodied carbon — primary material production (no recycling credit). DEFRA also publishes a 'closed-loop' value that nets out recycled content. This is for specific product forms (cans/foil), not raw material.",
                    "staleness_flag": False,
                })
                break

    # --- Rail ---
    val, uom = get_factor("Scope 3", "Business travel- land", "Rail", "National rail")
    if val:
        factors.append({
            "activity_id": "rail_intercity",
            "activity_name": "Rail (National)",
            "category": "ground_transport",
            "scope": "scope_3",
            "source": DEFRA_SOURCE,
            "factor": {
                "value_kg_co2e": round(val, 6),
                "unit_denominator": "passenger-km",
            },
            "original_value": {
                "value": val,
                "unit": "kg CO2e per passenger-km",
            },
            "geographic_scope": "UK",
            "staleness_flag": False,
        })

    # --- Bus ---
    # Scan for bus factor with kg CO2e
    for row_num in range(7, ws.max_row + 1):
        scope = str(_cell(ws, row_num, 2) or "").strip()
        l1 = str(_cell(ws, row_num, 3) or "").strip()
        l2 = str(_cell(ws, row_num, 4) or "").strip()
        l3 = str(_cell(ws, row_num, 5) or "").strip()
        ghg = str(_cell(ws, row_num, 9) or "").strip()
        val = _cell(ws, row_num, 10)
        if (scope == "Scope 3" and l1 == "Business travel- land" and l2 == "Bus"
            and "average" in l3.lower() and ghg == "kg CO2e"
            and val is not None and isinstance(val, (int, float))):
            factors.append({
                "activity_id": "bus",
                "activity_name": f"Bus ({l3})",
                "category": "ground_transport",
                "scope": "scope_3",
                "source": DEFRA_SOURCE,
                "factor": {
                    "value_kg_co2e": round(val, 6),
                    "unit_denominator": "passenger-km",
                },
                "original_value": {
                    "value": val,
                    "unit": "kg CO2e per passenger-km",
                },
                "geographic_scope": "UK",
                "staleness_flag": False,
            })
            break

    return factors


# ============================================================
# GHG PROTOCOL EXTRACTION
# ============================================================

def extract_ghg_protocol(wb):
    """Extract and normalize emission factors from GHG Protocol workbook."""
    factors = []

    # --- Natural Gas (Stationary Combustion sheet, row 44) ---
    # Row 44: col2='Natural gas', col3='Natural gas4', col4=48 (LHV TJ/Gg)
    # IMPORTANT: Row 8 has "Natural Gas Liquids" (NGL) — a DIFFERENT fuel.
    # Must match col2 exactly as "Natural gas" (not "Natural Gas Liquids").
    ws = wb["Stationary Combustion"]
    for row_num in range(2, ws.max_row + 1):
        col2 = str(_cell(ws, row_num, 2) or "").strip()
        if col2.lower() == "natural gas":  # exact match on col2, not col3
            co2_per_tj = _cell(ws, row_num, 5)  # kg CO2 / TJ (col 5)
            if co2_per_tj and isinstance(co2_per_tj, (int, float)):
                co2_per_kwh = co2_per_tj / TJ_TO_KWH
                factors.append({
                    "activity_id": "natural_gas_stationary",
                    "activity_name": "Natural Gas (Stationary Combustion)",
                    "category": "stationary_combustion",
                    "scope": "scope_1",
                    "source": {**GHG_PROTOCOL_SOURCE, "underlying_data_source": "IPCC 2006 Guidelines"},
                    "factor": {
                        "value_kg_co2e": round(co2_per_kwh, 6),
                        "unit_denominator": "kWh",
                        "co2_component": round(co2_per_kwh, 6),
                        "ch4_component": None,
                        "n2o_component": None,
                    },
                    "original_value": {
                        "value": co2_per_tj,
                        "unit": "kg CO2 per TJ",
                        "notes": "CO2 only. CH4/N2O in separate table. LHV (Net CV) basis."
                    },
                    "geographic_scope": "Global (IPCC default)",
                    "methodology_notes": "IPCC 2006 default factor. Lower Heating Value (Net CV) basis. CO2 only — does not include CH4/N2O from combustion.",
                    "staleness_flag": True,  # IPCC 2006 factors are 18+ years old
                })
                break

    # --- US Electricity (Electricity US sheet) ---
    # No US average row exists. eGRID subregions only. Col 2=name, col 3=CO2 lb/MWh
    # CH4 and N2O are in lb/GWh (not MWh!) — verified from header row 5
    # Skip for now — no clean US average available in this file.
    # (This is itself a friction point: GHG Protocol doesn't provide a US average)

    # --- Air Travel — US rows (Mobile Combustion - Public, rows 28-30) ---
    # Row 28: US, Air Travel - Short Haul, <300 miles, 0.207 kg/passenger-mile
    # Row 29: US, Air - Medium Haul, >=300 and <2300 miles, 0.129 kg/passenger-mile
    # Row 30: US, Air - Long Haul, >=2300 miles, 0.163 kg/passenger-mile
    ws = wb["Mobile Combustion - Public"]
    us_air_rows = [
        (28, "air_travel_short_haul", "Air Travel - Short Haul", "< 300 miles"),
        (29, "air_travel_medium_haul", "Air Travel - Medium Haul", "300-2300 miles"),
        (30, "air_travel_long_haul", "Air Travel - Long Haul", ">= 2300 miles"),
    ]
    for row, aid, aname, dist_note in us_air_rows:
        co2 = _cell(ws, row, 5)   # kg CO2 / passenger-mile
        ch4 = _cell(ws, row, 7)   # g CH4 / passenger-mile
        n2o = _cell(ws, row, 9)   # g N2O / passenger-mile
        if co2 and isinstance(co2, (int, float)):
            ch4_co2e = (ch4 / 1000) * GWP_CH4 if ch4 and isinstance(ch4, (int, float)) else 0
            n2o_co2e = (n2o / 1000) * GWP_N2O if n2o and isinstance(n2o, (int, float)) else 0
            total_per_mile = co2 + ch4_co2e + n2o_co2e
            total_per_km = total_per_mile / MILE_TO_KM

            factors.append({
                "activity_id": aid,
                "activity_name": aname,
                "category": "air_travel",
                "scope": "scope_3",
                "source": {**GHG_PROTOCOL_SOURCE, "underlying_data_source": "DEFRA (via EPA)"},
                "factor": {
                    "value_kg_co2e": round(total_per_km, 6),
                    "unit_denominator": "passenger-km",
                    "co2_component": round(co2 / MILE_TO_KM, 6),
                    "ch4_component": round(ch4_co2e / MILE_TO_KM, 6),
                    "n2o_component": round(n2o_co2e / MILE_TO_KM, 6),
                },
                "original_value": {
                    "value": co2,
                    "unit": "kg CO2 per passenger-mile",
                    "notes": f"US region. Distance: {dist_note}. Does NOT include radiative forcing."
                },
                "geographic_scope": "US",
                "methodology_notes": "GHG Protocol sources US air travel from DEFRA via EPA. Does NOT include radiative forcing. Converted from passenger-miles to passenger-km.",
                "staleness_flag": False,
            })

    # --- UK Air Travel (rows 7-16) — Average Passenger class only ---
    uk_air_rows = [
        (7, "air_travel_domestic", "Air Travel - Domestic", "UK domestic"),
        (8, "air_travel_short_haul", "Air Travel - Short Haul", "Up to 3700km"),
        (11, "air_travel_long_haul", "Air Travel - Long Haul", "Over 3700km"),
        (16, "air_travel_international", "Air Travel - International", "Non-UK international"),
    ]
    for row, aid, aname, dist_note in uk_air_rows:
        co2 = _cell(ws, row, 5)   # kg CO2 / passenger-km
        ch4 = _cell(ws, row, 7)   # g CH4 / passenger-km
        n2o = _cell(ws, row, 9)   # g N2O / passenger-km
        if co2 and isinstance(co2, (int, float)):
            ch4_co2e = (ch4 / 1000) * GWP_CH4 if ch4 and isinstance(ch4, (int, float)) else 0
            n2o_co2e = (n2o / 1000) * GWP_N2O if n2o and isinstance(n2o, (int, float)) else 0
            total = co2 + ch4_co2e + n2o_co2e

            factors.append({
                "activity_id": f"{aid}_uk_ghgp",
                "activity_name": f"{aname} (UK, GHG Protocol)",
                "category": "air_travel",
                "scope": "scope_3",
                "source": {**GHG_PROTOCOL_SOURCE, "underlying_data_source": "DEFRA 2022"},
                "factor": {
                    "value_kg_co2e": round(total, 6),
                    "unit_denominator": "passenger-km",
                },
                "original_value": {
                    "value": co2,
                    "unit": "kg CO2 per passenger-km",
                    "notes": f"UK region. {dist_note}. Average Passenger class. Does NOT include radiative forcing."
                },
                "geographic_scope": "UK",
                "methodology_notes": "From DEFRA 2022 WITHOUT radiative forcing. Compare to DEFRA 2025 flat file which INCLUDES RF — this explains the ~2x difference.",
                "staleness_flag": True,  # Uses DEFRA 2022, not 2025
            })

    # --- US Passenger Car (row 38) ---
    co2 = _cell(ws, 38, 5)  # kg CO2 / vehicle-mile
    ch4 = _cell(ws, 38, 7)  # g CH4 / vehicle-mile
    n2o = _cell(ws, 38, 9)  # g N2O / vehicle-mile
    if co2 and isinstance(co2, (int, float)):
        ch4_co2e = (ch4 / 1000) * GWP_CH4 if ch4 and isinstance(ch4, (int, float)) else 0
        n2o_co2e = (n2o / 1000) * GWP_N2O if n2o and isinstance(n2o, (int, float)) else 0
        total_per_mile = co2 + ch4_co2e + n2o_co2e
        total_per_km = total_per_mile / MILE_TO_KM

        factors.append({
            "activity_id": "passenger_car",
            "activity_name": "Passenger Car",
            "category": "ground_transport",
            "scope": "scope_1",
            "source": {**GHG_PROTOCOL_SOURCE, "underlying_data_source": "EPA"},
            "factor": {
                "value_kg_co2e": round(total_per_km, 6),
                "unit_denominator": "vehicle-km",
            },
            "original_value": {
                "value": co2,
                "unit": "kg CO2 per vehicle-mile",
                "notes": "US Passenger Car. Converted from vehicle-miles to vehicle-km."
            },
            "geographic_scope": "US",
            "methodology_notes": "From EPA via GHG Protocol. WARNING: GHG Protocol value (0.175 kg CO2/mi) differs significantly from EPA Hub value (0.297 kg CO2/mi). Likely different fleet definitions or data vintages. The EPA Hub 2025 value appears to be a newer, higher estimate. This 51% discrepancy is itself a key finding.",
            "staleness_flag": True,
        })

    # --- US Bus (row 36) ---
    co2 = _cell(ws, 36, 5)
    ch4 = _cell(ws, 36, 7)
    n2o = _cell(ws, 36, 9)
    if co2 and isinstance(co2, (int, float)):
        ch4_co2e = (ch4 / 1000) * GWP_CH4 if ch4 and isinstance(ch4, (int, float)) else 0
        n2o_co2e = (n2o / 1000) * GWP_N2O if n2o and isinstance(n2o, (int, float)) else 0
        total_per_mile = co2 + ch4_co2e + n2o_co2e
        total_per_km = total_per_mile / MILE_TO_KM

        factors.append({
            "activity_id": "bus",
            "activity_name": "Bus",
            "category": "ground_transport",
            "scope": "scope_3",
            "source": {**GHG_PROTOCOL_SOURCE, "underlying_data_source": "EPA"},
            "factor": {
                "value_kg_co2e": round(total_per_km, 6),
                "unit_denominator": "passenger-km",
            },
            "original_value": {
                "value": co2,
                "unit": "kg CO2 per passenger-mile",
            },
            "geographic_scope": "US",
            "staleness_flag": False,
        })

    # --- Freight (Mobile Combustion - Freight sheet) ---
    ws = wb["Mobile Combustion - Freight"]
    for row_num in range(4, ws.max_row + 1):
        vehicle = str(_cell(ws, row_num, 2) or "").strip()
        if "hgv" in vehicle.lower():
            co2_ef = _cell(ws, row_num, 6)
            ef_unit = str(_cell(ws, row_num, 7) or "").strip()
            if co2_ef and isinstance(co2_ef, (int, float)) and "tonne-kilometer" in ef_unit.lower():
                factors.append({
                    "activity_id": "freight_road_truck",
                    "activity_name": "Freight - Road (HGV)",
                    "category": "freight",
                    "scope": "scope_3",
                    "source": {**GHG_PROTOCOL_SOURCE, "underlying_data_source": "DEFRA"},
                    "factor": {
                        "value_kg_co2e": round(co2_ef, 6),
                        "unit_denominator": "tonne-km",
                    },
                    "original_value": {
                        "value": co2_ef,
                        "unit": ef_unit,
                    },
                    "geographic_scope": "UK",
                    "methodology_notes": "Sourced from DEFRA. Weight-distance basis (tonne-km).",
                    "staleness_flag": False,
                })
                break

    return factors


# ============================================================
# HELPERS
# ============================================================

def _cell(ws, row, col):
    """Safe cell value retrieval. Returns None for empty/merged cells.
    Returns numeric values as float, strings as strings."""
    try:
        cell = ws.cell(row=row, column=col)
        val = cell.value
        if val is None:
            return None
        if isinstance(val, (int, float)):
            return float(val)
        # Try to parse as float (handles string numbers)
        s = str(val).strip()
        try:
            return float(s)
        except (ValueError, TypeError):
            return s
    except Exception:
        return None


# ============================================================
# MAIN
# ============================================================

def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    all_factors = []

    # EPA
    print("Extracting EPA factors...")
    wb_epa = openpyxl.load_workbook(RAW_DIR / "epa-ghg-emission-factors-hub-2025.xlsx", data_only=True)
    epa_factors = extract_epa(wb_epa)
    print(f"  Found {len(epa_factors)} factors")
    all_factors.extend(epa_factors)

    # DEFRA
    print("Extracting DEFRA factors...")
    wb_defra = openpyxl.load_workbook(RAW_DIR / "defra-ghg-conversion-factors-2025-flat.xlsx", data_only=True)
    defra_factors = extract_defra(wb_defra)
    print(f"  Found {len(defra_factors)} factors")
    all_factors.extend(defra_factors)

    # GHG Protocol
    print("Extracting GHG Protocol factors...")
    wb_ghg = openpyxl.load_workbook(RAW_DIR / "ghg-protocol-emission-factors-v2.0.xlsx", data_only=True)
    ghg_factors = extract_ghg_protocol(wb_ghg)
    print(f"  Found {len(ghg_factors)} factors")
    all_factors.extend(ghg_factors)

    # Write individual source files
    with open(OUT_DIR / "epa_factors.json", "w") as f:
        json.dump(epa_factors, f, indent=2)

    with open(OUT_DIR / "defra_factors.json", "w") as f:
        json.dump(defra_factors, f, indent=2)

    with open(OUT_DIR / "ghg_protocol_factors.json", "w") as f:
        json.dump(ghg_factors, f, indent=2)

    # Write combined file (source factors only — no country data)
    with open(OUT_DIR / "all_factors.json", "w") as f:
        json.dump(all_factors, f, indent=2)

    # Load and write country-specific factors separately
    country_factors_path = Path(__file__).parent.parent / "data" / "country_factors.json"
    if country_factors_path.exists():
        print("Loading country-specific factors...")
        with open(country_factors_path) as f:
            country_factors = json.load(f)
        print(f"  Found {len(country_factors)} country-specific factors")

        with open(OUT_DIR / "country_factors.json", "w") as f:
            json.dump(country_factors, f, indent=2)

    # Write comparison view grouped by activity_id
    comparison = {}
    for factor in all_factors:
        aid = factor["activity_id"]
        if aid not in comparison:
            comparison[aid] = {
                "activity_name": factor["activity_name"],
                "category": factor["category"],
                "normalized_unit": factor["factor"]["unit_denominator"],
                "sources": [],
            }
        comparison[aid]["sources"].append({
            "source_name": factor["source"]["name"],
            "value_kg_co2e": factor["factor"]["value_kg_co2e"],
            "geographic_scope": factor["geographic_scope"],
            "original_value": factor["original_value"]["value"],
            "original_unit": factor["original_value"]["unit"],
            "methodology_notes": factor.get("methodology_notes", ""),
            "staleness_flag": factor.get("staleness_flag", False),
        })

    # Add variance calculation
    for aid, data in comparison.items():
        values = [s["value_kg_co2e"] for s in data["sources"] if s["value_kg_co2e"]]
        if len(values) >= 2:
            min_v, max_v = min(values), max(values)
            avg = sum(values) / len(values)
            variance_pct = ((max_v - min_v) / avg) * 100 if avg > 0 else 0
            data["variance_pct"] = round(variance_pct, 1)
            data["min_value"] = min_v
            data["max_value"] = max_v
        else:
            data["variance_pct"] = None

    with open(OUT_DIR / "comparison.json", "w") as f:
        json.dump(comparison, f, indent=2)

    # Print summary
    print(f"\nTotal factors extracted: {len(all_factors)}")
    print(f"\nComparison summary:")
    print(f"{'Activity':<40} {'Sources':>7} {'Variance':>10} {'Unit':>15}")
    print("-" * 75)
    for aid, data in sorted(comparison.items()):
        n_sources = len(data["sources"])
        var_str = f"{data['variance_pct']}%" if data["variance_pct"] is not None else "N/A"
        print(f"{data['activity_name']:<40} {n_sources:>7} {var_str:>10} {data['normalized_unit']:>15}")


if __name__ == "__main__":
    main()
